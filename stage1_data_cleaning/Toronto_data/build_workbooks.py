#!/usr/bin/env python3
"""Build and validate the two stage 1–3 Excel deliverables with openpyxl."""

from __future__ import annotations

import math
import textwrap
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parent
INVENTORY_PATH = ROOT / "indicator_inventory.xlsx"
DICTIONARY_PATH = ROOT / "data_dictionary.xlsx"
PREVIEW_DIR = Path("/private/tmp/toronto_xlsx_previews")

SOURCE_URL = "https://www12.statcan.gc.ca/census-recensement/2021/dp-pd/prof/details/download-telecharger.cfm?DGUIDlist=2021A000210&GENDERlist=1&HEADERlist=0&Lang=E&STATISTIClist=1%2C4"
LICENSE_URL = "https://www.statcan.gc.ca/en/terms-conditions/open-licence"
DOWNLOAD_FILE = "98-401-X2021007_eng_CSV.zip"
MISSING_NOTE = "Blank, ..., .., suppression and not-applicable values remain missing; never replace with 0."

INDICATORS = [
    dict(field="population_2021", cid=1, name="Population, 2021", description="2021 total population", column="C1_COUNT_TOTAL", unit="persons", value_format="integer", source_year=2021, sample="100% data", non_missing=6240, missing=7, notes="Official CT-level count."),
    dict(field="population_density_km2", cid=6, name="Population density per square kilometre", description="Population per square kilometre", column="C1_COUNT_TOTAL", unit="persons/km²", value_format="decimal", source_year=2021, sample="100% data", non_missing=6240, missing=7, notes="Official CT-level density."),
    dict(field="age_65_plus_pct", cid=37, name="65 years and over", description="Share of population aged 65 years and over", column="C10_RATE_TOTAL", unit="percent", value_format="0–100 percent", source_year=2021, sample="100% data", non_missing=6166, missing=81, notes="17.8 means 17.8%, not 0.178."),
    dict(field="average_household_size", cid=57, name="Average household size", description="Average persons per household", column="C1_COUNT_TOTAL", unit="persons per household", value_format="decimal", source_year=2021, sample="100% data", non_missing=6161, missing=86, notes="Official source name is Average household size."),
    dict(field="median_total_income_2020", cid=113, name="Median total income in 2020 among recipients ($)", description="Median 2020 total income among recipients", column="C1_COUNT_TOTAL", unit="CAD", value_format="currency/count", source_year=2020, sample="100% data", non_missing=6100, missing=147, notes="Reference year is 2020; published in the 2021 Census Profile using 2021 CT geography."),
    dict(field="low_income_lim_at_pct", cid=345, name="Prevalence of low income based on the Low-income measure, after tax (LIM-AT) (%)", description="Prevalence of after-tax low income under LIM-AT", column="C10_RATE_TOTAL", unit="percent", value_format="0–100 percent", source_year=2020, sample="100% data", non_missing=6100, missing=147, notes="Reference year is 2020; published in the 2021 Census Profile."),
    dict(field="renter_households_pct", cid=1416, name="Renter", description="Share of households that rent their dwelling", column="C10_RATE_TOTAL", unit="percent", value_format="0–100 percent", source_year=2021, sample="25% sample data", non_missing=6158, missing=89, notes="Official CT-level long-form sample indicator."),
    dict(field="shelter_cost_30pct_plus_pct", cid=1467, name="Spending 30% or more of income on shelter costs", description="Share spending at least 30% of income on shelter", column="C10_RATE_TOTAL", unit="percent", value_format="0–100 percent", source_year=2021, sample="25% sample data", non_missing=6099, missing=148, notes="Official CT-level long-form sample indicator."),
    dict(field="bachelor_or_higher_age25_64_pct", cid=2024, name="Bachelor's degree or higher", description="Share of persons aged 25–64 with a bachelor's degree or higher", column="C10_RATE_TOTAL", unit="percent", value_format="0–100 percent", source_year=2021, sample="25% sample data", non_missing=6158, missing=89, notes="Uses ID 2024 for persons aged 25–64; do not substitute ID 2008."),
    dict(field="unemployment_rate_pct", cid=2230, name="Unemployment rate", description="Unemployment rate", column="C10_RATE_TOTAL", unit="percent", value_format="0–100 percent", source_year=2021, sample="25% sample data", non_missing=6158, missing=89, notes="Official CT-level long-form sample indicator."),
]


HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Aptos", size=10, color="1F2933")
THIN_GRAY = Side(style="thin", color="D9E2F3")


def style_table_sheet(ws, table_name: str, widths: list[float]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 42
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="0F243E"))
    for row in ws.iter_rows(min_row=2):
        max_len = 1
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)
            max_len = max(max_len, len(str(cell.value or "")))
        ws.row_dimensions[row[0].row].height = min(72, max(24, 24 + 12 * (max_len // 70)))
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = width
    table = Table(displayName=table_name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.auto_filter.ref = ws.dimensions


def inventory_rows():
    rows = []
    for item in INDICATORS:
        rows.append([
            item["field"], item["cid"], item["name"], item["description"], item["column"], item["unit"], item["value_format"],
            "Census Tract", "Census Tract", "direct_join", "DGUID", item["source_year"], 2021, item["sample"], "available",
            item["non_missing"], item["missing"], item["non_missing"] / 6247, MISSING_NOTE, "Statistics Canada", SOURCE_URL,
            DOWNLOAD_FILE, "2026-07-13", "Statistics Canada Open Licence", item["notes"],
        ])
    return rows


def build_inventory() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "indicator_inventory"
    headers = [
        "standardized_field", "characteristic_id", "official_indicator_name", "description", "data_column", "unit", "value_format",
        "boundary_unit", "indicator_unit", "join_method", "join_id", "source_year", "product_year", "sample_type", "quality_flag",
        "non_missing_count", "missing_count", "coverage_rate", "missing_value_note", "source_organization", "source_url", "download_file",
        "access_date", "license", "notes",
    ]
    ws.append(headers)
    for row in inventory_rows():
        ws.append(row)
    widths = [36, 22, 48, 40, 22, 22, 20, 21, 21, 18, 13, 16, 16, 22, 18, 22, 18, 18, 48, 24, 72, 36, 16, 32, 56]
    style_table_sheet(ws, "IndicatorInventoryTable", widths)
    for cell in ws["R"][1:]:
        cell.number_format = "0.0%"
        cell.alignment = Alignment(horizontal="right", vertical="top")
    for col in ("B", "L", "M"):
        for cell in ws[col][1:]:
            cell.number_format = "0"
    for col in ("P", "Q"):
        for cell in ws[col][1:]:
            cell.number_format = "#,##0"
    for cell in ws["U"][1:]:
        cell.hyperlink = SOURCE_URL
        cell.style = "Hyperlink"
        cell.alignment = Alignment(vertical="top", wrap_text=True, shrink_to_fit=True)

    source = wb.create_sheet("source_information")
    source.append(["field", "value", "notes"])
    source_rows = [
        ("country", "Canada", "Research country"),
        ("city", "Toronto", "Target city; the current data are not yet filtered to the city"),
        ("data_product", "Census Profile, 2021 Census of Population", "Official indicator product"),
        ("source_organization", "Statistics Canada", "Official publisher"),
        ("source_url", SOURCE_URL, "Stable official selection/download page"),
        ("download_file", DOWNLOAD_FILE, "Actual downloaded ZIP file name"),
        ("download_method", "Official page > comprehensive download > CMAs, tracted CAs and CTs > CSV > English", "Direct file URL is dynamically generated and was not retained as a stable link"),
        ("download_file_size", "249,720,408 bytes", "Local source ZIP size"),
        ("internal_csv", "98-401-X2021007_English_CSV_data.csv (2,634,464,532 bytes)", "Streamed from ZIP; do not open in Excel"),
        ("product_year", "2021", "Census Profile and CT geography vintage"),
        ("access_date", "2026-07-13", "Recorded download/access date"),
        ("license", "Statistics Canada Open Licence", LICENSE_URL),
        ("original_geographic_level", "Census tract", "All ten selected indicators are officially published at CT level"),
        ("current_processed_coverage", "Canada-wide Census Tract dataset (6,247 CTs)", "Not a Toronto City subset"),
        ("next_stage", "Use the Toronto City boundary to derive an in-city DGUID list, then filter and join", "Do not treat Toronto CMA code 535 as Toronto City"),
        ("primary_join_id", "DGUID (text)", "No concatenation required"),
        ("alternate_id", "ALT_GEO_CODE in indicators / CTUID in boundary (text)", "Preserve leading zeros and formatting"),
    ]
    for row in source_rows:
        source.append(row)
    style_table_sheet(source, "SourceInformationTable", [30, 100, 80])
    for row in range(2, source.max_row + 1):
        if source.cell(row, 1).value == "source_url":
            source.cell(row, 2).hyperlink = SOURCE_URL
            source.cell(row, 2).style = "Hyperlink"
            source.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True, shrink_to_fit=True)
        if source.cell(row, 1).value == "license":
            source.cell(row, 3).hyperlink = LICENSE_URL
            source.cell(row, 3).style = "Hyperlink"

    quality = wb.create_sheet("quality_summary")
    quality.append(["item_type", "metric_or_indicator", "value", "missing_count", "coverage_rate", "notes"])
    quality_rows = [
        ["dataset", "total_ct_count", 6247, None, None, "Canada-wide Census Tracts"],
        ["dataset", "unique_dguid_count", 6247, None, None, "DGUID is unique"],
        ["dataset", "duplicate_dguid_count", 0, None, None, "No duplicate DGUID rows"],
    ]
    for item in INDICATORS:
        quality_rows.append(["indicator", item["field"], item["non_missing"], item["missing"], item["non_missing"] / 6247, item["sample"]])
    quality_rows.extend([
        ["method_note", "100% vs 25% sample", "See notes", None, None, "100% data generally have no sampling error; 25% long-form indicators have sampling uncertainty. Both may have suppression/not-applicable missing values."],
        ["year_note", "cross-year indicators", "2020 reference", None, None, "Median income and LIM-AT reference 2020 but are published in the 2021 Census Profile on 2021 CT geography."],
        ["scope_note", "current coverage", "Canada-wide", None, None, "Not yet filtered to Toronto City; code 535 identifies Toronto CMA only."],
        ["handoff_note", "next stage", "Toronto City clipping", None, None, "Use a Toronto City boundary to derive the city DGUID list before joining or mapping."],
    ])
    for row in quality_rows:
        quality.append(row)
    style_table_sheet(quality, "QualitySummaryTable", [18, 38, 24, 16, 16, 86])
    for cell in quality["E"][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.0%"
    for col in ("C", "D"):
        for cell in quality[col][1:]:
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"

    wb.properties.title = "Toronto project — 2021 Census Tract indicator inventory"
    wb.properties.creator = "OpenAI Codex"
    wb.properties.subject = "Stage 1–3 Canada-wide CT indicator metadata"
    wb.save(INVENTORY_PATH)


def dictionary_rows():
    rows = [
        ["CENSUS_YEAR", "text", "Census year for the geography and profile record", "CENSUS_YEAR", None, "year", "2021", 2021, "not applicable", "Must be present; preserve as text", "metadata", "available", "Current output contains 2021 records."],
        ["DGUID", "text", "Dissemination Geography Unique Identifier", "DGUID", None, "identifier", "valid Statistics Canada DGUID", 2021, "not applicable", "Must be present; never coerce to a number", "primary join key", "available", "Preferred direct join field between boundary and indicator data."],
        ["ALT_GEO_CODE", "text", "Alternative geography code; corresponds to boundary CTUID at CT level", "ALT_GEO_CODE", None, "identifier", "valid CTUID format", 2021, "not applicable", "Must be present; preserve leading zeros and decimal formatting", "alternate join/check key", "available", "Treat as text."],
        ["GEO_LEVEL", "text", "Official geographic level", "GEO_LEVEL", None, "category", "Census tract", 2021, "not applicable", "Must equal Census tract", "scale validation", "available", "All output rows are Census tract."],
        ["GEO_NAME", "text", "Official geography name", "GEO_NAME", None, "text", "official value", 2021, "not applicable", "Must be present", "label", "available", "Not a substitute for DGUID."],
        ["DATA_QUALITY_FLAG", "text", "Statistics Canada five-digit geography data-quality flag", "DATA_QUALITY_FLAG", None, "flag", "official flag code", 2021, "not applicable", "Blank remains missing", "quality metadata", "available", "Retained from the source profile."],
    ]
    chinese_descriptions = {
        "population_2021": "2021 total population",
        "population_density_km2": "Population density per square kilometre",
        "age_65_plus_pct": "Share aged 65 years and over",
        "average_household_size": "Average household size",
        "median_total_income_2020": "Median 2020 total income among recipients",
        "low_income_lim_at_pct": "After-tax LIM-AT low-income prevalence",
        "renter_households_pct": "Share of renter households",
        "shelter_cost_30pct_plus_pct": "Share spending 30% or more of income on shelter",
        "bachelor_or_higher_age25_64_pct": "Share aged 25–64 with bachelor's degree or higher",
        "unemployment_rate_pct": "Unemployment rate",
    }
    for item in INDICATORS:
        is_percent = item["unit"] == "percent"
        valid = "0–100; e.g. 17.8 means 17.8%" if is_percent else ">= 0"
        rows.append([
            item["field"], "number", chinese_descriptions[item["field"]], item["name"], item["cid"], item["unit"], valid,
            item["source_year"], item["sample"], MISSING_NOTE, "direct_join attribute (DGUID)", "available",
            item["notes"] + " Current CSV is one row per Canada-wide CT, not a Toronto City subset.",
        ])
    return rows


def build_dictionary() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "data_dictionary"
    headers = ["field_name", "field_type", "description", "official_source_name", "characteristic_id", "unit", "valid_range", "source_year", "sample_type", "missing_value_rule", "join_role", "quality_flag", "notes"]
    ws.append(headers)
    for row in dictionary_rows():
        ws.append(row)
    widths = [48, 16, 46, 54, 22, 24, 30, 16, 20, 54, 28, 18, 70]
    style_table_sheet(ws, "DataDictionaryTable", widths)
    for col in ("E", "H"):
        for cell in ws[col][1:]:
            if isinstance(cell.value, int):
                cell.number_format = "0"
    wb.properties.title = "Toronto project — processed CT data dictionary"
    wb.properties.creator = "OpenAI Codex"
    wb.properties.subject = "Fields in canada_ct_10_indicators_2021.csv"
    wb.save(DICTIONARY_PATH)


def display_value(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, float) and cell.number_format == "0.0%":
        return f"{value:.1%}"
    if isinstance(value, int) and cell.number_format == "#,##0":
        return f"{value:,}"
    return str(value)


def get_font(size=18, bold=False):
    names = [
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf" if bold else "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/HelveticaNeue.ttc",
    ]
    for name in names:
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def render_range(ws, start_col: int, end_col: int, out_path: Path) -> None:
    header_font = get_font(18, True)
    body_font = get_font(17)
    title_font = get_font(24, True)
    cols = list(range(start_col, end_col + 1))
    px_widths = [max(110, min(430, int((ws.column_dimensions[ws.cell(1, c).column_letter].width or 13) * 7.2))) for c in cols]
    wrapped = []
    row_heights = []
    for r in range(1, ws.max_row + 1):
        row_lines = []
        max_lines = 1
        for c, width in zip(cols, px_widths):
            text = display_value(ws.cell(r, c))
            chars = max(8, int(width / (10 if r == 1 else 9)))
            lines = textwrap.wrap(text, width=chars, break_long_words=True, break_on_hyphens=True) or [""]
            row_lines.append(lines)
            max_lines = max(max_lines, len(lines))
        wrapped.append(row_lines)
        row_heights.append(max(42 if r == 1 else 32, 12 + max_lines * (22 if r == 1 else 21)))
    title_h = 58
    width = sum(px_widths) + 2
    height = title_h + sum(row_heights) + 2
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, width, title_h), fill="#123B5D")
    draw.text((14, 15), f"{ws.title} — columns {ws.cell(1,start_col).column_letter}:{ws.cell(1,end_col).column_letter}", font=title_font, fill="white")
    y = title_h
    for r, (row_lines, row_h) in enumerate(zip(wrapped, row_heights), 1):
        x = 1
        fill = "#2F75B5" if r == 1 else ("#F3F7FB" if r % 2 == 0 else "#FFFFFF")
        fg = "#FFFFFF" if r == 1 else "#1F2933"
        font = header_font if r == 1 else body_font
        for lines, col_w in zip(row_lines, px_widths):
            draw.rectangle((x, y, x + col_w, y + row_h), fill=fill, outline="#D9E3F0", width=1)
            ty = y + 8
            for line in lines:
                draw.text((x + 8, ty), line, font=font, fill=fg)
                ty += 22 if r == 1 else 21
            x += col_w
        y += row_h
    out_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(out_path, "PNG", optimize=True)


def validate_and_render() -> None:
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    inv = load_workbook(INVENTORY_PATH, data_only=False)
    assert inv.sheetnames == ["indicator_inventory", "source_information", "quality_summary"]
    assert inv["indicator_inventory"].max_row == 11 and inv["indicator_inventory"].max_column == 25
    assert inv["indicator_inventory"].freeze_panes == "A2"
    assert len(inv["indicator_inventory"].tables) == 1
    assert all(inv["indicator_inventory"].cell(r, 18).number_format == "0.0%" for r in range(2, 12))
    render_range(inv["indicator_inventory"], 1, 13, PREVIEW_DIR / "indicator_inventory_A_M.png")
    render_range(inv["indicator_inventory"], 14, 25, PREVIEW_DIR / "indicator_inventory_N_Y.png")
    render_range(inv["source_information"], 1, 3, PREVIEW_DIR / "source_information.png")
    render_range(inv["quality_summary"], 1, 6, PREVIEW_DIR / "quality_summary.png")

    dd = load_workbook(DICTIONARY_PATH, data_only=False)
    assert dd.sheetnames == ["data_dictionary"]
    assert dd["data_dictionary"].max_row == 17 and dd["data_dictionary"].max_column == 13
    assert dd["data_dictionary"].freeze_panes == "A2"
    assert len(dd["data_dictionary"].tables) == 1
    render_range(dd["data_dictionary"], 1, 7, PREVIEW_DIR / "data_dictionary_A_G.png")
    render_range(dd["data_dictionary"], 8, 13, PREVIEW_DIR / "data_dictionary_H_M.png")
    print({
        "inventory_sheets": inv.sheetnames,
        "inventory_rows": inv["indicator_inventory"].max_row - 1,
        "dictionary_fields": dd["data_dictionary"].max_row - 1,
        "preview_dir": str(PREVIEW_DIR),
    })


if __name__ == "__main__":
    build_inventory()
    build_dictionary()
    validate_and_render()
